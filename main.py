from tkinter import *
import openpyxl

root=Tk()
root.title("Valentin, MercadoLibre Inventory")
root.geometry("1000x500")

#update the Listbox
def update(data):
    #clear the listbox
    my_list.delete(0, END)

    #Add products to my_list
    for item in data:
        my_list.insert(END, item)

#Update entry box with listbox clicked
def fillout(e): #event
    #Delete whatever is in the entry box
    my_entry.delete(0, END)
    #Add clicked list item to entry box
    content=my_list.get(ACTIVE)
    my_entry.insert(0, content)

    #Delete whatever is in the box_entry
    box_entry.delete(0, END)
    #I searched for the box according to the reference
    box=""
    for row in range(2,number_rows+1):
        for col in range(1,2):
            if excel_worksheet.cell(row,col).value == content:
                box=str(int(excel_worksheet.cell(row,col+1).value))
    box_entry.insert(0, box)

#create function to check entry vs listbox
def check(e):
    # grab what was typed
    typed= my_entry.get()
    if typed == "":
        data = products
    else:
        data=[]
        for item in products:
            if typed.lower() in item.lower():
                data.append(item)
    #update our listbos with selected items
    update(data)

def change_box():
    new_box_number=0
    try:
        new_box_number=int(box_entry.get())
    except:
        print("Exception")
    box_entry.delete(0,END)
    reference=my_list.get(ACTIVE)


    for row in range(2,number_rows):
        for col in range(1,2):
            if excel_worksheet.cell(row,col).value == reference:
                excel_worksheet.cell(row,col+1).value=new_box_number
    """
    num=0
    for row in range(2,number_rows+1):
        for col in range(1,2):
            excel_worksheet.cell(row,col+1).value=num
            #num+=1
    """
    excel_workbook.save(path)

#create a label
my_label= Label(root, text="Start typing...",
                font=("Helvetica",14), fg="black")
my_label.pack(pady=20, side=TOP)

#create an entry box
my_entry = Entry(root, font=("Helvetica",15), width=50)
my_entry.pack()

change_button=Button(root,text="change box", command=change_box)
change_button.pack(pady=20, side=RIGHT)

box_label = Label(root, text="BOX",
                    font=("Helvetica",15), fg="black")
box_label.pack(pady=20, side=RIGHT)

box_entry = Entry(root, font=("Helvetica",15), width=5)
box_entry.pack(side=RIGHT)

# Create a listbox
my_list = Listbox(root, font=("Helvetica",10), width=110, height=200)
my_list.pack(pady=40)

#Create a list
products=[]

#path="Publications.xlsx"
path="Publications.xlsx"
excel_workbook = openpyxl.load_workbook(path)
excel_worksheet = excel_workbook["Python"]

number_cols=excel_worksheet.max_column
number_rows=excel_worksheet.max_row
print("cols :"+ str(number_cols) + ", rows: " + str(number_rows))

for rowi in range(2,number_rows+1):
    for col in range(1,2):
        if excel_worksheet.cell(rowi,col).value !=None:
            products.append(excel_worksheet.cell(rowi,col).value)

update(products)

#Create a binding on the listbox onclick
my_list.bind("<<ListboxSelect>>",fillout)

#create a binding on the entry box
my_entry.bind("<KeyRelease>", check)

root.mainloop()
print("END OF LINE")
